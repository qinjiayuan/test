# !/usr/bin python3
# encoding: utf-8 -*-
# @file     : file_load.py
# @author   : Gowent 
# @Time     : 2025/2/22 20:49
# @Copyright: Personal
import openpyxl
import yaml


def read_excel(filepath,sheet_name):
    #获取整个文档对象
    wb= openpyxl.load_workbook(filepath)
    #获取某个sheet工作表的数据
    sheet_data = wb[sheet_name]
    lines_count = sheet_data.max_row
    cols_count = sheet_data.max_column
    test_data = []
    for row in range(2,lines_count+1):#这里的行就是肉眼所见的行，不需要用1开始，因为1是表头
        row_data = []
        for col in range(1,cols_count+1):
            cell_data = sheet_data.cell(row,col).value
            row_data.append(cell_data)
        test_data.append(row_data)
    print(test_data)

    return test_data


def load_yaml_file(filepath):
    with open(file=filepath,mode='r',encoding='utf-8') as f :
        content = yaml.load(f,Loader=yaml.FullLoader)
        print(content)
        return content

def write_yaml(filepath,content):
    with open(file=filepath,mode='w',encoding='utf-8') as f :
        yaml.dump(content,f,Dumper=yaml.Dumper)

if __name__ == '__main__':
    # read_excel(r"../data/data.xlsx","立即购买接口")
    load_yaml_file("../data/data.yml")